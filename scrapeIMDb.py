import openpyxl
import strip as strip
from bs4 import BeautifulSoup
from openpyxl import* ##?????????????????????
import  requests
#creatig new file excel
excel=openpyxl.Workbook()
#print(excel.sheetnames)
#to change the active sheet name
sheet =excel.active
sheet.title='top Rated movies'
print(excel.sheetnames)
sheet.append(['Rank Movie','Movie Name','Year Of Realese','Ratint IMDb'])


#url of the website
try:
  source=requests.get('https://www.imdb.com/chart/top/')
  source.raise_for_status()
  soup=BeautifulSoup(source.text,'html.parser')
  #print(soup)
  movies=soup.find('tbody',{'class','lister-list'}).findAll('tr')
  #print(len(movies))
  for movie in movies:#loop on each movie and it's details
      name=movie.find('td',{'class','titleColumn'}).a.text
      rank= (movie.find('td', {'class', 'titleColumn'}).text).strip().split('.')[0]#????['1', '\n      The Shawshank Redemption\n(1994)']
      year=movie.find('td', {'class', 'titleColumn'}).span.text.strip('()')
      ratting=movie.find('td',{'class','ratingColumn imdbRating'}).strong.text
      print(rank,year,name,ratting)
      sheet.append([rank,year,name,ratting])

except Exception as e:
    print(e)
excel.save('TOP Movie Rating.xlsx')